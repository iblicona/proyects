import sqlite3
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, send_file

# 🔹 Excel bonito
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# 🔹 Crear base de datos
def crear_bd():
    conn = sqlite3.connect("biblioteca.db")
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        apellidos TEXT NOT NULL,
        matricula TEXT NOT NULL UNIQUE,
        correo TEXT NOT NULL,
        tipo TEXT NOT NULL,
        area TEXT NOT NULL
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS libros (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titulo TEXT NOT NULL,
        autor TEXT NOT NULL,
        editorial TEXT,
        anio INTEGER,
        isbn TEXT UNIQUE,
        cantidad INTEGER NOT NULL
    )
    """)

    conn.commit()
    conn.close()

# 🔹 Página principal
@app.route("/")
def inicio():
    return render_template("inicio.html")

# 🔹 Registrar usuarios
@app.route("/registro", methods=["POST"])
def registrar():
    nombre = request.form["nombre"]
    apellidos = request.form["apellidos"]
    matricula = request.form["matricula"]
    correo = request.form["correo"]
    tipo = request.form["tipo"]
    area = request.form["area"]

    try:
        conn = sqlite3.connect("biblioteca.db")
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO usuarios (nombre, apellidos, matricula, correo, tipo, area)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (nombre, apellidos, matricula, correo, tipo, area))

        conn.commit()
        conn.close()

        return redirect(url_for("inicio"))

    except sqlite3.IntegrityError:
        return "La matrícula ya está registrada"

# 🔹 Registrar libros
@app.route("/registrar_libro", methods=["POST"])
def registrar_libro():
    titulo = request.form["titulo"]
    autor = request.form["autor"]
    editorial = request.form["editorial"]
    anio = request.form["anio"]
    isbn = request.form["isbn"]
    cantidad = request.form["cantidad"]

    try:
        conn = sqlite3.connect("biblioteca.db")
        cursor = conn.cursor()

        cursor.execute("""
            INSERT INTO libros (titulo, autor, editorial, anio, isbn, cantidad)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (titulo, autor, editorial, anio, isbn, cantidad))

        conn.commit()
        conn.close()

        return redirect(url_for("inicio"))

    except sqlite3.IntegrityError:
        return "Ese ISBN ya está registrado"

# 🔹 Exportar usuarios a Excel bonito
@app.route("/exportar_usuarios")
def exportar_usuarios():
    conn = sqlite3.connect("biblioteca.db")

    df = pd.read_sql_query("""
        SELECT 
            id AS ID,
            nombre AS Nombre,
            apellidos AS Apellidos,
            matricula AS Matrícula,
            correo AS Correo,
            tipo AS Tipo,
            area AS Área
        FROM usuarios
        ORDER BY nombre ASC
    """, conn)

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # 🔹 Título
    ws.merge_cells("A1:G1")
    titulo = ws["A1"]
    titulo.value = "REPORTE DE USUARIOS - BIBLIOTECA"
    titulo.font = Font(size=14, bold=True)
    titulo.alignment = Alignment(horizontal="center")

    # 🔹 Subtítulo
    ws.merge_cells("A2:G2")
    ws["A2"].value = "Generado automáticamente por el sistema"
    ws["A2"].alignment = Alignment(horizontal="center")

    # 🔹 Insertar datos
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 🔹 Estilo encabezados
    header_fill = PatternFill(start_color="6b3e26", end_color="6b3e26", fill_type="solid")

    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 🔹 Ajustar ancho automático (sin error de celdas combinadas)
    for i, column_cells in enumerate(ws.iter_cols(min_row=3, max_row=ws.max_row, max_col=7), 1):
        max_length = 0
        column_letter = get_column_letter(i)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 4

    # 🔹 Filtros
    ws.auto_filter.ref = "A3:G3"

    archivo = "usuarios.xlsx"
    wb.save(archivo)

    return send_file(archivo, as_attachment=True)

if __name__ == "__main__":
    crear_bd()
    app.run(debug=True, host="0.0.0.0", port=5001)
